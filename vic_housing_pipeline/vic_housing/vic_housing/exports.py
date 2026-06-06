"""
exports.py — Export pipeline data to CSV files and an Excel dashboard.

Outputs (written to ./exports/):
  sales_medians.csv
  rental_medians.csv
  building_approvals.csv
  lending_rates.csv
  asx_announcements.csv
  vic_housing_dashboard.xlsx   ← multi-sheet Excel with a derived Yields sheet
"""

from __future__ import annotations

import sqlite3
from pathlib import Path

from .core import DB_PATH, BASE_DIR, get_conn, get_logger

log     = get_logger("exports")
OUT_DIR = BASE_DIR / "exports"
OUT_DIR.mkdir(exist_ok=True)

TABLES = [
    "sales_medians",
    "rental_medians",
    "building_approvals",
    "lending_rates",
    "asx_announcements",
]


# ---------------------------------------------------------------------------
# CSV export
# ---------------------------------------------------------------------------
def export_csvs() -> None:
    import csv
    conn = get_conn()
    for table in TABLES:
        path = OUT_DIR / f"{table}.csv"
        try:
            cursor = conn.execute(f"SELECT * FROM {table}")
            cols   = [d[0] for d in cursor.description]
            rows   = cursor.fetchall()
            with open(path, "w", newline="", encoding="utf-8") as f:
                writer = csv.writer(f)
                writer.writerow(cols)
                writer.writerows(rows)
            log.info(f"  CSV: {path.name} ({len(rows)} rows)")
        except Exception as e:
            log.warning(f"  CSV export failed for {table}: {e}")
    conn.close()


# ---------------------------------------------------------------------------
# Excel dashboard
# ---------------------------------------------------------------------------
def _df_from_table(table: str):
    import pandas as pd
    try:
        conn = get_conn()
        df   = pd.read_sql_query(f"SELECT * FROM {table}", conn)
        conn.close()
        return df
    except Exception as e:
        log.warning(f"  Could not read {table}: {e}")
        import pandas as pd
        return pd.DataFrame()


def _build_yields_sheet(sales_df, rental_df):
    """
    Join sales_medians (houses) to rental_medians (all dwellings) on
    suburb + lga + period to compute gross yield % = (annual_rent / price) * 100.
    """
    import pandas as pd

    if sales_df.empty or rental_df.empty:
        return pd.DataFrame(columns=[
            "period", "suburb", "lga", "median_price",
            "median_weekly_rent", "gross_yield_pct",
        ])

    s = sales_df[sales_df["dwelling_type"] == "house"][
        ["period", "suburb", "lga", "median_price"]
    ].copy()
    r = rental_df[rental_df["dwelling_type"] == "all"][
        ["period", "suburb", "lga", "median_rent"]
    ].copy()

    merged = s.merge(r, on=["period", "suburb", "lga"], how="inner")
    merged = merged.dropna(subset=["median_price", "median_rent"])
    merged = merged[merged["median_price"] > 0]

    merged["gross_yield_pct"] = (
        merged["median_rent"] * 52 / merged["median_price"] * 100
    ).round(2)

    merged = merged.rename(columns={"median_rent": "median_weekly_rent"})
    return merged.sort_values(["period", "gross_yield_pct"], ascending=[False, False])


def export_excel() -> None:
    try:
        import pandas as pd
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils.dataframe import dataframe_to_rows
    except ImportError as e:
        log.warning(f"Excel export requires pandas and openpyxl: {e}")
        return

    path = OUT_DIR / "vic_housing_dashboard.xlsx"

    dfs   = {t: _df_from_table(t) for t in TABLES}
    yields = _build_yields_sheet(dfs["sales_medians"], dfs["rental_medians"])

    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    HEADER_FILL  = PatternFill("solid", fgColor="1F3864")
    HEADER_FONT  = Font(bold=True, color="FFFFFF", size=10)
    ALT_FILL     = PatternFill("solid", fgColor="EEF2F7")

    def write_sheet(wb, name: str, df, freeze_row: int = 1):
        ws = wb.create_sheet(title=name[:31])
        if df.empty:
            ws.cell(1, 1, f"No data — run the pipeline first")
            return ws

        # Header
        for col_idx, col_name in enumerate(df.columns, 1):
            cell = ws.cell(1, col_idx, col_name)
            cell.font      = HEADER_FONT
            cell.fill      = HEADER_FILL
            cell.alignment = Alignment(horizontal="center")

        # Data rows
        for row_idx, row in enumerate(df.itertuples(index=False), 2):
            fill = ALT_FILL if row_idx % 2 == 0 else None
            for col_idx, val in enumerate(row, 1):
                cell = ws.cell(row_idx, col_idx, val)
                if fill:
                    cell.fill = fill

        # Auto-width (capped at 40)
        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

        ws.freeze_panes = ws.cell(freeze_row + 1, 1)
        ws.auto_filter.ref = ws.dimensions
        return ws

    sheet_map = {
        "Sales_Medians":    dfs["sales_medians"],
        "Rental_Medians":   dfs["rental_medians"],
        "Building_Approvl": dfs["building_approvals"],
        "Lending_Rates":    dfs["lending_rates"],
        "ASX_Announcemnts": dfs["asx_announcements"],
        "Yields_Calc":      yields,
    }

    for sheet_name, df in sheet_map.items():
        write_sheet(wb, sheet_name, df)
        log.info(f"  Sheet '{sheet_name}': {len(df)} rows")

    wb.save(path)
    log.info(f"Excel dashboard saved: {path}")


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------
def run() -> None:
    log.info("Exporting CSVs...")
    export_csvs()
    log.info("Building Excel dashboard...")
    export_excel()
    log.info("Export complete.")
