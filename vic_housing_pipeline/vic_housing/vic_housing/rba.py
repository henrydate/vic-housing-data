"""
rba.py -- RBA Statistical Tables connector.

Files (XLSX, no auth required):
  F5  f05hist.xlsx  Indicator Lending Rates
  F6  f06hist.xlsx  Housing Lending Rates (OO vs Investor / P&I vs IO)  <- key table
  F7  f07hist.xlsx  Business Lending Rates (opt-in via RBA_INCLUDE_F7=true)

RBA XLSX structure (confirmed):
  Row 0:  Table title  (e.g. "F5  INDICATOR LENDING RATES")
  Row 1:  "Title"       + series titles
  Row 2:  "Description" + series descriptions      <- use for labels
  Row 3:  "Frequency"
  Row 4:  "Type"
  Row 5:  "Units"
  Row 6-7: blank
  Row 8:  "Source"
  Row 9:  "Publication date"
  Row 10: "Series ID"  + actual series IDs         <- dynamically detected
  Row 11+: data rows   (col 0 = datetime object, cols 1+ = float values)
"""

from __future__ import annotations

import datetime
import os
import re
from io import BytesIO

from .core import build_session, get_logger, upsert

log = get_logger("rba")

RBA_XLS_BASE = "https://www.rba.gov.au/statistics/tables/xls"

SERIES = {
    "F5": {
        "url":     f"{RBA_XLS_BASE}/f05hist.xlsx",
        "label":   "Indicator Lending Rates",
        "primary": True,
    },
    "F6": {
        "url":     f"{RBA_XLS_BASE}/f06hist.xlsx",
        "label":   "Housing Lending Rates (OO vs Investor / P&I vs IO)",
        "primary": True,
    },
    "F7": {
        "url":     f"{RBA_XLS_BASE}/f07hist.xlsx",
        "label":   "Business Lending Rates (by firm size)",
        "primary": False,  # set RBA_INCLUDE_F7=true to enable
    },
}

INCLUDE_F7 = os.getenv("RBA_INCLUDE_F7", "false").lower() == "true"


def _parse_rba_xlsx(content: bytes, table_id: str) -> list[dict]:
    """
    Parse an RBA historical XLSX file.
    Dynamically locates the 'Series ID' row and 'Description' row
    rather than assuming fixed row numbers.
    Handles datetime objects in the date column.
    """
    from openpyxl import load_workbook

    rows_out = []
    try:
        wb = load_workbook(BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
        all_rows = list(ws.iter_rows(values_only=True))

        if len(all_rows) < 5:
            log.warning(f"  {table_id}: too few rows ({len(all_rows)})")
            return []

        # -- Locate key rows ---------------------------------------------------
        series_id_row_idx   = None
        description_row_idx = None

        for i, row in enumerate(all_rows):
            if not row or row[0] is None:
                continue
            cell0 = str(row[0]).strip().strip("'").lower()
            if cell0 == "series id":
                series_id_row_idx = i
            elif cell0 in ("description", "title") and description_row_idx is None:
                description_row_idx = i

        if series_id_row_idx is None:
            log.warning(f"  {table_id}: 'Series ID' row not found")
            return []

        series_ids   = [str(v or "").strip().strip("'") for v in all_rows[series_id_row_idx]]
        descriptions = (
            [str(v or "").strip().strip("'") for v in all_rows[description_row_idx]]
            if description_row_idx is not None
            else [""] * len(series_ids)
        )

        month_map = {
            "Jan": "01", "Feb": "02", "Mar": "03", "Apr": "04",
            "May": "05", "Jun": "06", "Jul": "07", "Aug": "08",
            "Sep": "09", "Oct": "10", "Nov": "11", "Dec": "12",
        }

        # -- Parse data rows (everything after the Series ID row) --------------
        for row in all_rows[series_id_row_idx + 1:]:
            if not row or row[0] is None:
                continue

            date_val = row[0]

            # openpyxl returns date cells as datetime objects
            if isinstance(date_val, (datetime.datetime, datetime.date)):
                period = date_val.strftime("%Y-%m")
            else:
                raw = str(date_val).strip().strip("'")
                m = re.match(r"(\d{4})-([A-Za-z]{3})", raw)
                if m:
                    period = f"{m.group(1)}-{month_map.get(m.group(2), '01')}"
                else:
                    m2 = re.match(r"(\d{4})-(\d{2})", raw)
                    if m2:
                        period = f"{m2.group(1)}-{m2.group(2)}"
                    else:
                        continue  # not a date row

            for col_idx in range(1, len(row)):
                if col_idx >= len(series_ids):
                    break
                series_id = series_ids[col_idx]
                if not series_id or series_id in ("None", ""):
                    continue

                label = descriptions[col_idx] if col_idx < len(descriptions) else ""
                val   = row[col_idx]

                if val is None or str(val).strip() in ("", "..", "na", "NA"):
                    continue
                try:
                    rate = float(val)
                except (ValueError, TypeError):
                    continue

                rows_out.append({
                    "period":       period,
                    "series_id":    f"{table_id}/{series_id}",
                    "series_label": label or f"{table_id}/{series_id}",
                    "rate_pct":     rate,
                })

    except Exception as e:
        log.warning(f"  {table_id}: XLSX parse error: {e}")

    return rows_out


def run() -> int:
    session  = build_session()
    all_rows = []

    for table_id, meta in SERIES.items():
        if not meta["primary"] and not INCLUDE_F7:
            log.info(f"Skipping {table_id} ({meta['label']}) -- set RBA_INCLUDE_F7=true to enable")
            continue

        log.info(f"Fetching RBA {table_id}: {meta['label']}")
        try:
            r = session.get(meta["url"], timeout=60)
            r.raise_for_status()
        except Exception as e:
            log.warning(f"  {table_id}: fetch failed: {e}")
            continue

        rows = _parse_rba_xlsx(r.content, table_id)
        log.info(f"  {table_id}: {len(rows)} observations")
        all_rows.extend(rows)

    new = upsert("lending_rates", all_rows, ["period", "series_id"])
    log.info(f"RBA: {len(all_rows)} rows -> {new} new inserted")
    return new
