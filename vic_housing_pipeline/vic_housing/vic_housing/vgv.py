"""
vgv.py — Valuer-General Victoria connector.

Source: data.vic.gov.au CKAN API
Dataset: "Victorian Property Sales Statistics" (median sale prices by suburb/LGA)
Cadence: Quarterly + Annual releases

Strategy:
  1. Query CKAN package search for the VGV sales dataset.
  2. For each XLSX resource, download and parse via openpyxl.
  3. Detect period, LGA, suburb, and dwelling type from file/sheet metadata.
  4. Upsert into sales_medians.
"""

from __future__ import annotations

import re
from io import BytesIO

import openpyxl

from .core import build_session, cached_get, get_logger, upsert

log = get_logger("vgv")

CKAN_BASE = "https://discover.data.vic.gov.au/api/3/action"

# Known dataset IDs for VGV median property sales
VGV_PACKAGE_IDS = [
    "victorian-property-sales-statistics-median-house-by-suburb",
    "victorian-property-sales-statistics-median-unit-by-suburb",
    "victorian-property-sales-statistics-all-dwellings-by-suburb",
]

DWELLING_KEYWORDS = {
    "house": "house",
    "houses": "house",
    "unit": "unit",
    "units": "unit",
    "apartment": "unit",
    "land": "land",
    "vacant": "land",
}


def _detect_dwelling_type(text: str) -> str:
    t = text.lower()
    for kw, dtype in DWELLING_KEYWORDS.items():
        if kw in t:
            return dtype
    return "house"  # default


def _detect_period(text: str) -> str | None:
    """Extract YYYY-QN or YYYY from filename / sheet name."""
    m = re.search(r"(\d{4})[-_ ]?q(\d)", text, re.I)
    if m:
        return f"{m.group(1)}-Q{m.group(2)}"
    m = re.search(r"(\d{4})", text)
    if m:
        return m.group(1)
    return None


def _find_header_row(ws) -> int | None:
    """Return row index (1-based) where the data header lives."""
    for row in ws.iter_rows(max_row=20):
        for cell in row:
            v = str(cell.value or "").lower()
            if "suburb" in v or "lga" in v or "median" in v:
                return cell.row
    return None


def _parse_sheet(ws, period: str, dwelling_type: str) -> list[dict]:
    header_row = _find_header_row(ws)
    if header_row is None:
        log.debug(f"  No header found in sheet '{ws.title}'")
        return []

    headers = [str(ws.cell(header_row, c).value or "").lower().strip()
               for c in range(1, ws.max_column + 1)]

    # Map column names to indices
    def col(keyword: str) -> int | None:
        for i, h in enumerate(headers):
            if keyword in h:
                return i
        return None

    suburb_col = col("suburb")
    lga_col    = col("lga") or col("local government")
    median_col = col("median")
    sales_col  = col("number") or col("sales") or col("count")

    if median_col is None:
        log.debug(f"  No median column in sheet '{ws.title}'")
        return []

    rows = []
    for r in range(header_row + 1, ws.max_row + 1):
        suburb = ws.cell(r, (suburb_col or 0) + 1).value if suburb_col is not None else None
        lga    = ws.cell(r, (lga_col or 0) + 1).value    if lga_col    is not None else None
        median = ws.cell(r, median_col + 1).value
        sales  = ws.cell(r, (sales_col or 0) + 1).value  if sales_col  is not None else None

        if median is None:
            continue
        try:
            median = float(str(median).replace(",", "").replace("$", "").strip())
        except ValueError:
            continue

        rows.append({
            "period":        period,
            "lga":           str(lga).strip()    if lga    else None,
            "suburb":        str(suburb).strip() if suburb else None,
            "dwelling_type": dwelling_type,
            "median_price":  median,
            "num_sales":     int(sales) if sales and str(sales).strip().isdigit() else None,
        })
    return rows


def _process_resource(session, resource: dict) -> list[dict]:
    url  = resource.get("url", "")
    name = resource.get("name", "") or resource.get("description", "")

    if not url.lower().endswith((".xlsx", ".xls")):
        return []

    period       = _detect_period(url) or _detect_period(name) or "unknown"
    dwelling_type = _detect_dwelling_type(url + " " + name)

    log.info(f"  Downloading: {url} → period={period} type={dwelling_type}")
    resp = cached_get(session, url)
    if resp is None:
        return []

    try:
        wb = openpyxl.load_workbook(BytesIO(resp.content), read_only=True, data_only=True)
    except Exception as e:
        log.warning(f"  Could not open workbook {url}: {e}")
        return []

    all_rows = []
    for ws in wb.worksheets:
        # Some files have one sheet per dwelling type — re-detect
        sheet_type = _detect_dwelling_type(ws.title) if ws.title else dwelling_type
        sheet_period = _detect_period(ws.title) or period
        rows = _parse_sheet(ws, sheet_period, sheet_type)
        all_rows.extend(rows)
        log.debug(f"    Sheet '{ws.title}': {len(rows)} rows")

    return all_rows


def run() -> int:
    """Fetch VGV data and upsert into sales_medians. Returns new row count."""
    session   = build_session()
    all_rows  = []

    for pkg_id in VGV_PACKAGE_IDS:
        log.info(f"Querying CKAN package: {pkg_id}")
        resp = cached_get(session, f"{CKAN_BASE}/package_show", params={"id": pkg_id})
        if resp is None:
            log.warning(f"  Could not fetch package metadata for {pkg_id}")
            continue

        try:
            data      = resp.json()
            resources = data["result"]["resources"]
        except Exception as e:
            log.warning(f"  Malformed CKAN response for {pkg_id}: {e}")
            continue

        log.info(f"  Found {len(resources)} resources")
        for resource in resources:
            rows = _process_resource(session, resource)
            all_rows.extend(rows)

    new = upsert("sales_medians", all_rows, ["period", "lga", "suburb", "dwelling_type"])
    log.info(f"VGV: {len(all_rows)} rows parsed → {new} new inserted")
    return new
