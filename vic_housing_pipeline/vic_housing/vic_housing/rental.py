"""
rental.py — DFFH Rental Report connector (RTBA bond data).

Source: data.vic.gov.au — Department of Families, Fairness & Housing
Dataset: "Rental Report — Quarterly data" (median weekly rents by suburb/LGA)
Cadence: Quarterly

Strategy:
  1. Query CKAN for the RTBA rental report package.
  2. Download each XLSX resource.
  3. Parse median weekly rent by suburb, dwelling type, and quarter.
  4. Upsert into rental_medians.
"""

from __future__ import annotations

import re
from io import BytesIO

import openpyxl

from .core import build_session, cached_get, get_logger, upsert

log = get_logger("rental")

CKAN_BASE = "https://discover.data.vic.gov.au/api/3/action"

RENTAL_PACKAGE_IDS = [
    "rental-report-quarterly-moving-annual-rents-by-suburb",
    "rental-report-quarterly-bond-data-suburb",
]

BEDROOM_MAP = {
    "1 bedroom": "1br", "1br": "1br", "one bedroom": "1br",
    "2 bedroom": "2br", "2br": "2br", "two bedroom": "2br",
    "3 bedroom": "3br", "3br": "3br", "three bedroom": "3br",
    "4 bedroom": "4br", "4br": "4br",
    "all":       "all", "total": "all",
}


def _detect_period(text: str) -> str | None:
    m = re.search(r"(\d{4})[-_ ]?q(\d)", text, re.I)
    if m:
        return f"{m.group(1)}-Q{m.group(2)}"
    m = re.search(r"(\d{4})", text)
    if m:
        return m.group(1)
    return None


def _detect_bedroom_type(text: str) -> str:
    t = text.lower()
    for kw, btype in BEDROOM_MAP.items():
        if kw in t:
            return btype
    return "all"


def _find_header_row(ws) -> int | None:
    for row in ws.iter_rows(max_row=20):
        for cell in row:
            v = str(cell.value or "").lower()
            if "suburb" in v or "lga" in v or "median" in v or "rent" in v:
                return cell.row
    return None


def _parse_sheet(ws, period: str, dwelling_type: str) -> list[dict]:
    header_row = _find_header_row(ws)
    if header_row is None:
        return []

    headers = [str(ws.cell(header_row, c).value or "").lower().strip()
               for c in range(1, ws.max_column + 1)]

    def col(keyword: str) -> int | None:
        for i, h in enumerate(headers):
            if keyword in h:
                return i
        return None

    suburb_col = col("suburb")
    lga_col    = col("lga") or col("local government")
    median_col = col("median") or col("rent")

    if median_col is None:
        return []

    rows = []
    for r in range(header_row + 1, ws.max_row + 1):
        suburb = ws.cell(r, (suburb_col or 0) + 1).value if suburb_col is not None else None
        lga    = ws.cell(r, (lga_col or 0) + 1).value    if lga_col    is not None else None
        median = ws.cell(r, median_col + 1).value

        if median is None:
            continue
        try:
            median = float(str(median).replace(",", "").replace("$", "").strip())
        except ValueError:
            continue

        rows.append({
            "period":        period,
            "lga":           str(lga).strip()    if lga    else None,
            "suburb":        str(suburb).strip() if suburb else None,
            "dwelling_type": dwelling_type,
            "median_rent":   median,
        })
    return rows


def _process_resource(session, resource: dict) -> list[dict]:
    url  = resource.get("url", "")
    name = resource.get("name", "") or resource.get("description", "")

    if not url.lower().endswith((".xlsx", ".xls")):
        return []

    period        = _detect_period(url) or _detect_period(name) or "unknown"
    dwelling_type = _detect_bedroom_type(url + " " + name)

    log.info(f"  Downloading: {url} → period={period} type={dwelling_type}")
    resp = cached_get(session, url)
    if resp is None:
        return []

    try:
        wb = openpyxl.load_workbook(BytesIO(resp.content), read_only=True, data_only=True)
    except Exception as e:
        log.warning(f"  Could not open workbook {url}: {e}")
        return []

    all_rows = []
    for ws in wb.worksheets:
        sheet_type   = _detect_bedroom_type(ws.title) if ws.title else dwelling_type
        sheet_period = _detect_period(ws.title) or period
        rows         = _parse_sheet(ws, sheet_period, sheet_type)
        all_rows.extend(rows)

    return all_rows


def run() -> int:
    session  = build_session()
    all_rows = []

    for pkg_id in RENTAL_PACKAGE_IDS:
        log.info(f"Querying CKAN package: {pkg_id}")
        resp = cached_get(session, f"{CKAN_BASE}/package_show", params={"id": pkg_id})
        if resp is None:
            log.warning(f"  Could not fetch package for {pkg_id}")
            continue

        try:
            data      = resp.json()
            resources = data["result"]["resources"]
        except Exception as e:
            log.warning(f"  Malformed response for {pkg_id}: {e}")
            continue

        for resource in resources:
            rows = _process_resource(session, resource)
            all_rows.extend(rows)

    new = upsert("rental_medians", all_rows, ["period", "lga", "suburb", "dwelling_type"])
    log.info(f"Rental: {len(all_rows)} rows parsed → {new} new inserted")
    return new
