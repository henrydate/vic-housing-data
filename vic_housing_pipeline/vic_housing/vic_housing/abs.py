"""
abs.py -- ABS Building Approvals connector.

Endpoint (post 29 Nov 2024 migration):
  data.api.abs.gov.au/rest/data/ABS,BA_GCCSA,1.0.0/all
  Requires Accept: application/vnd.sdmx.data+json  (otherwise returns XML)

Dataflow: BA_GCCSA  (Building Approvals by Greater Capital City Statistical Area)
  - Monthly cadence, original + seasonally adjusted
  - Dimensions: MEASURE, REGION, DWELLING_STRUCTURE, TSEST, TIME_PERIOD

XLSX fallback: abs.gov.au latest-release page
"""

from __future__ import annotations

import json
import re
from io import BytesIO

from .core import build_session, get_logger, upsert

log = get_logger("abs")

ABS_API_BASE = "https://data.api.abs.gov.au/rest/data"
DATAFLOW_ID  = "ABS,BA_GCCSA,1.0.0"
START_PERIOD = "2015-01"

# Region codes to keep (VIC + national total)
# 2GMEL = Greater Melbourne, AUS = Australia total
KEEP_REGIONS = {"greater melbourne", "australia", "victoria", "rest of vic", "rest of victoria"}

ABS_LATEST_PAGE = (
    "https://www.abs.gov.au/statistics/industry/building-and-construction/"
    "building-approvals-australia/latest-release"
)

# JSON accept header -- without this the API returns XML which we cannot parse
SDMX_JSON_ACCEPT = "application/vnd.sdmx.data+json;version=1.0"


def _fetch_sdmx_json(session, url: str, params: dict) -> dict | None:
    """Stream the SDMX JSON response (can be large) and return parsed dict."""
    try:
        r = session.get(
            url, params=params, timeout=120,
            headers={"Accept": SDMX_JSON_ACCEPT},
            stream=True,
        )
        r.raise_for_status()
        chunks = []
        for chunk in r.iter_content(chunk_size=65536):
            chunks.append(chunk)
        return json.loads(b"".join(chunks))
    except Exception as e:
        log.warning(f"  SDMX fetch error: {e}")
        return None


def _parse_sdmx(data: dict) -> list[dict]:
    """
    ABS SDMX-JSON v1.0 grouped format:
      data.dataSets[0].series  -- keyed by series dim indices (e.g. '0:0:0:0:0:0:0:0')
      each series has .observations keyed by time-period index (e.g. '0', '1', ...)
      data.structure.dimensions.series  -- series-level dimensions
      data.structure.dimensions.observation -- TIME_PERIOD values list
    """
    rows = []
    try:
        structure    = data["data"]["structure"]
        series_dims  = structure["dimensions"]["series"]       # list of dim dicts
        obs_dims     = structure["dimensions"]["observation"]  # [TIME_PERIOD]

        # Index helpers
        def dim_idx(dims, dim_id):
            for i, d in enumerate(dims):
                if d["id"] == dim_id:
                    return i
            return None

        # Keep only the dwelling-COUNT measure + residential building types.
        DWELLING_MEASURE = "number of dwelling units"
        DWELLING_TYPES = {
            "houses": "house",
            "dwellings excluding houses": "unit",
            "total residential": "total",
        }
        PIN = {"VALUE": "total", "SECTOR": "total sectors", "WORK_TYPE": "total work"}
        measure_idx  = dim_idx(series_dims, "MEASURE")
        region_idx   = dim_idx(series_dims, "REGION")
        btype_idx    = dim_idx(series_dims, "BUILDING_TYPE")
        tsest_idx    = dim_idx(series_dims, "TSEST")
        pin_idx      = {k: dim_idx(series_dims, k) for k in PIN}
        time_values  = obs_dims[0]["values"]

        for series_key, series_data in data["data"]["dataSets"][0]["series"].items():
            keys = series_key.split(":")

            # 1) dwelling COUNT only + pin sector/value/work-type to aggregates
            if measure_idx is not None:
                mname = series_dims[measure_idx]["values"][int(keys[measure_idx])]["name"].lower()
                if mname != DWELLING_MEASURE:
                    continue
            if any(pin_idx[k] is not None and
                   series_dims[pin_idx[k]]["values"][int(keys[pin_idx[k]])]["name"].lower() != v
                   for k, v in PIN.items()):
                continue

            # 2) this state's regions + national only
            region = (series_dims[region_idx]["values"][int(keys[region_idx])]["name"]
                      if region_idx is not None else "Australia")
            if region.lower() not in KEEP_REGIONS:
                continue

            # 3) residential dwelling types only -> house / unit / total
            btype = (series_dims[btype_idx]["values"][int(keys[btype_idx])]["name"].lower()
                     if btype_idx is not None else "")
            dtype = DWELLING_TYPES.get(btype)
            if dtype is None:
                continue

            seasonality = "original"
            if tsest_idx is not None:
                ts = series_dims[tsest_idx]["values"][int(keys[tsest_idx])]["name"].lower()
                seasonality = ("seasonally_adjusted" if "season" in ts
                               else "trend" if "trend" in ts else "original")

            for time_key, obs_vals in series_data.get("observations", {}).items():
                t_idx = int(time_key)
                if t_idx >= len(time_values):
                    continue
                value = obs_vals[0] if obs_vals else None
                if value is None:
                    continue
                rows.append({
                    "period":        time_values[t_idx]["id"],
                    "region":        region,
                    "dwelling_type": dtype,
                    "seasonality":   seasonality,
                    "num_approvals": int(round(float(value))),
                    "value_000":     None,
                })
    except Exception as e:
        log.warning(f"SDMX parse error: {e}")
    return rows


def _find_xlsx_on_page(session) -> str | None:
    """Scrape the ABS latest-release page to find a current XLSX download URL."""
    try:
        from bs4 import BeautifulSoup
        r = session.get(ABS_LATEST_PAGE, timeout=30)
        r.raise_for_status()
        soup = BeautifulSoup(r.content, "lxml")
        for a in soup.find_all("a", href=True):
            href = a["href"]
            if re.search(r"8731\d*do\d+.*\.xlsx", href, re.I) or (
                "building-approvals" in href.lower() and href.endswith(".xlsx")
            ):
                return href if href.startswith("http") else "https://www.abs.gov.au" + href
    except Exception as e:
        log.warning(f"  Could not scrape ABS page: {e}")
    return None


def _parse_xlsx_fallback(content: bytes) -> list[dict]:
    import openpyxl
    rows = []
    try:
        wb = openpyxl.load_workbook(BytesIO(content), read_only=True, data_only=True)
        for ws in wb.worksheets:
            header_row = None
            for r in ws.iter_rows(max_row=30):
                for cell in r:
                    if any(kw in str(cell.value or "").lower() for kw in ("month", "period", "date")):
                        header_row = cell.row
                        break
                if header_row:
                    break
            if not header_row:
                continue

            for r in range(header_row + 1, (ws.max_row or 0) + 1):
                period_raw = ws.cell(r, 1).value
                val        = ws.cell(r, 2).value
                if not period_raw or not val:
                    continue
                m = re.search(r"(\d{4})[-/](\d{1,2})", str(period_raw))
                if not m:
                    continue
                try:
                    rows.append({
                        "period":        f"{m.group(1)}-{int(m.group(2)):02d}",
                        "region":        "Australia",
                        "dwelling_type": "total",
                        "seasonality":   "original",
                        "num_approvals": int(float(str(val).replace(",", ""))),
                        "value_000":     None,
                    })
                except ValueError:
                    continue
    except Exception as e:
        log.warning(f"XLSX fallback parse error: {e}")
    return rows


def run() -> int:
    session = build_session()
    rows    = []

    # -- Primary: SDMX JSON API (requires Accept header) -------------------
    url = f"{ABS_API_BASE}/{DATAFLOW_ID}/all"
    log.info(f"Fetching ABS BA_GCCSA: {url}")
    data = _fetch_sdmx_json(session, url, {
        "startPeriod":            START_PERIOD,
        "detail":                 "Full",
        "dimensionAtObservation": "TIME_PERIOD",
    })
    if data:
        rows = _parse_sdmx(data)
        log.info(f"  SDMX: {len(rows)} observations parsed")

    # -- Fallback: XLSX from latest-release page ---------------------------
    if not rows:
        log.info("Falling back to ABS XLSX download")
        xlsx_url = _find_xlsx_on_page(session)
        if xlsx_url:
            log.info(f"  Found XLSX: {xlsx_url}")
            try:
                r = session.get(xlsx_url, timeout=60)
                r.raise_for_status()
                rows = _parse_xlsx_fallback(r.content)
                log.info(f"  XLSX fallback: {len(rows)} rows")
            except Exception as e:
                log.warning(f"  XLSX download failed: {e}")
        else:
            log.warning("  Could not find XLSX URL on ABS page")

    new = upsert("building_approvals", rows,
                 ["period", "region", "dwelling_type", "seasonality"])
    log.info(f"ABS: {len(rows)} rows -> {new} new inserted")
    return new
